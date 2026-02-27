"""
step2_filter.py
===============
Websedge Conference Lead Pipeline — Step 2: Filter & Flag
----------------------------------------------------------
Reads the enriched CSV from step1 and applies exclusion rules and flags.

HARD REMOVES (lead excluded entirely):
  - decisionMakerScore = 1  (residents, students, junior staff — no budget authority)
  - FinalScore < 20         (too low priority)
  - Institution in DO NOT CONTACT tab
  - Pharma company (keyword + known company list)
  - Public / government hospital (state hospitals, county hospitals, VA)

FLAGS ONLY (lead kept, flagged for your review):
  - PRIVATE_NONPROFIT     — private non-profit hospital (check university tab manually)
  - UNIVERSITY_HOSPITAL   — name contains "University" + hospital type
  - APPR_CONFLICT         — institution already being worked by a team member (APPR tab)
  - LOW_CONFIDENCE        — AI confidence = "low" (scores less reliable)
  - SAME_NAME_DIFF_ORG    — from CRM check (not applied here, added manually if needed)

Outputs:
  {output}_review.xlsx     ← formatted Excel ready for you to sort/filter and pick leads
  {output}_review.csv      ← same data as CSV backup
  {output}_excluded.csv    ← everything removed + reason

Usage:
  python step2_filter.py --input apa2026_extra_enriched.csv --output apa2026_extra
                         --xlsx "APA 2026.xlsx"

The _review.xlsx is colour-coded:
  Green  = clean, no flags
  Yellow = flagged (review before contacting)
  Grey   = would normally be excluded but kept for awareness (e.g. very low score)
"""

import argparse, json, os, re
import pandas as pd

# ── Reference data paths ───────────────────────────────────────────────────────
DEFAULT_XLSX = "/Users/benvarvill/Downloads/MRA Media work/APA 2026.xlsx"
BASE         = "/Users/benvarvill/Downloads/MRA Media work"

# ── Pharma filter ──────────────────────────────────────────────────────────────
PHARMA_KNOWN = {
    "boehringer ingelheim", "bristol myers squibb", "bristol-myers squibb",
    "janssen", "johnson & johnson", "johnson and johnson",
    "intra-cellular therapies", "neurocrine biosciences", "neurocrine",
    "otsuka pharmaceutical", "otsuka", "sage therapeutics",
    "takeda", "biogen", "indivior", "alkermes",
    "abbvie", "pfizer", "eli lilly", "astrazeneca", "novartis",
    "roche", "merck", "lundbeck", "sunovion", "acadia", "axsome",
}
PHARMA_KEYWORDS = [
    "pharmaceutical", "pharma", " biosciences", "biopharma", " therapeutics", "biologics",
]

def is_pharma(inst: str) -> bool:
    il = inst.lower()
    return any(k in il for k in PHARMA_KNOWN) or any(k in il for k in PHARMA_KEYWORDS)


# ── Public hospital filter ─────────────────────────────────────────────────────
PUBLIC_KNOWN = {
    "georgia regional hospital", "twin valley behavioral healthcare",
    "southwest connecticut mental health", "chicago-read mental health center",
    "austin state hospital", "metrohealth medical center",
    "metrohealth and case western", "norman regional health system",
}
PUBLIC_KEYWORDS = [
    "state hospital", "state psychiatric", "state mental health",
    "county hospital", "county mental health", "veterans affairs",
    "va medical center", " vamc",
]

def is_public_hospital(inst: str, inst_type: str) -> bool:
    if inst_type not in ("hospital", "unknown"):
        return False
    il = inst.lower()
    return any(k in il for k in PUBLIC_KNOWN) or any(k in il for k in PUBLIC_KEYWORDS)


# ── Private non-profit flag ────────────────────────────────────────────────────
def flag_private_nonprofit(inst: str, inst_type: str) -> bool:
    """Flag hospitals that appear to be private non-profits for manual review."""
    if inst_type != "hospital":
        return False
    il = inst.lower()
    # University in name = likely university-affiliated
    if re.search(r"\b(university|college of medicine|school of medicine|academic medical)\b", il):
        return True
    return False


# ── Reference data loaders ─────────────────────────────────────────────────────
def load_dnc(xlsx_path: str) -> set:
    """Load DO NOT CONTACT institution names (lowercase)."""
    try:
        df = pd.read_excel(xlsx_path, sheet_name="DO NOT CONTACT", header=None)
        names = set()
        for val in df[0].dropna():
            names.add(str(val).strip().lower())
        return names
    except Exception as e:
        print(f"  ⚠ Could not load DO NOT CONTACT tab: {e}")
        return set()


def load_key_contacts(xlsx_path: str) -> set:
    """
    Load APA Key Contacts (speakers, board members, officers etc.)
    Returns a set of normalised full names for fast lookup.
    Strips credentials like M.D., Ph.D. before storing.
    """
    try:
        df = pd.read_excel(xlsx_path, sheet_name="APA Key Contacts")
        names = set()
        cred_pat = re.compile(
            r",?\s+(M\.?D\.?|Ph\.?D\.?|D\.?O\.?|M\.?S\.?|M\.?B\.?A\.?|'?FAPA|DFAPA).*",
            re.IGNORECASE,
        )
        for val in df["Name"].dropna():
            clean = cred_pat.sub("", str(val)).strip().lower()
            if clean:
                names.add(clean)
        return names
    except Exception as e:
        print(f"  ⚠ Could not load APA Key Contacts tab: {e}")
        return set()


def is_key_contact(name: str, key_contacts: set) -> bool:
    """True if this person's name appears in the APA Key Contacts list."""
    if not key_contacts:
        return False
    cred_pat = re.compile(
        r",?\s+(M\.?D\.?|Ph\.?D\.?|D\.?O\.?|M\.?S\.?|M\.?B\.?A\.?|'?FAPA|DFAPA).*",
        re.IGNORECASE,
    )
    n = cred_pat.sub("", str(name)).strip().lower()
    return n in key_contacts


def load_appr_orgs(xlsx_path: str) -> set:
    """Load organisations already being worked by the team (APPR tab)."""
    try:
        df = pd.read_excel(xlsx_path, sheet_name="APPR")
        orgs = set()
        for col in ["Organisation", "Organization", "organization", "org", "Org"]:
            if col in df.columns:
                for val in df[col].dropna():
                    orgs.add(str(val).strip().lower())
                break
        return orgs
    except Exception as e:
        print(f"  ⚠ Could not load APPR tab: {e}")
        return set()


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).lower().strip())


def is_dnc(inst: str, dnc_set: set) -> bool:
    il = _norm(inst)
    for dnc in dnc_set:
        if dnc in il or il in dnc:
            return True
    return False


def is_appr_conflict(inst: str, appr_orgs: set) -> bool:
    il = _norm(inst)
    for org in appr_orgs:
        if not org or len(org) < 4:
            continue
        if org in il or il in org:
            return True
    return False


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Websedge Step 2 — Filter & Flag")
    parser.add_argument("--input",  required=True,           help="Enriched CSV from step1")
    parser.add_argument("--output", default="filtered",      help="Output filename prefix")
    parser.add_argument("--xlsx",   default=DEFAULT_XLSX,    help="Path to APA 20XX.xlsx")
    args = parser.parse_args()

    print(f"\n📂 Loading {args.input}...")
    df = pd.read_csv(args.input)
    print(f"   {len(df)} records loaded")

    print(f"📋 Loading reference data from {args.xlsx}...")
    dnc_set      = load_dnc(args.xlsx)
    appr_orgs    = load_appr_orgs(args.xlsx)
    key_contacts = load_key_contacts(args.xlsx)
    print(f"   {len(dnc_set)} DO NOT CONTACT | {len(appr_orgs)} APPR orgs | {len(key_contacts)} APA Key Contacts")

    included = []
    excluded = []

    for _, row in df.iterrows():
        inst      = str(row.get("currentInstitution") or row.get("Institution", ""))
        inst_type = str(row.get("institutionType", "unknown")).lower()
        dm_score  = int(row.get("decisionMakerScore", 0) or 0)
        score     = float(row.get("FinalScore", 0) or 0)
        confidence= str(row.get("confidence", "low")).lower()

        flags = []

        # ── Hard removes ──────────────────────────────────────────────────────
        if dm_score <= 1:
            excluded.append({**row, "excludeReason": "RESIDENT_OR_STUDENT (DM score ≤ 1)"})
            continue

        if score < 20:
            excluded.append({**row, "excludeReason": f"LOW_SCORE ({score:.1f} < 20)"})
            continue

        if is_dnc(inst, dnc_set):
            excluded.append({**row, "excludeReason": "DO_NOT_CONTACT"})
            continue

        if is_pharma(inst):
            excluded.append({**row, "excludeReason": "PHARMA"})
            continue

        if is_public_hospital(inst, inst_type):
            excluded.append({**row, "excludeReason": "PUBLIC_HOSPITAL"})
            continue

        # ── Flags (kept but highlighted) ──────────────────────────────────────
        if flag_private_nonprofit(inst, inst_type):
            flags.append("UNIVERSITY_HOSPITAL?")

        if inst_type == "hospital" and not flag_private_nonprofit(inst, inst_type):
            flags.append("PRIVATE_NONPROFIT — check university tab")

        if is_appr_conflict(inst, appr_orgs):
            flags.append("APPR_CONFLICT — team already working this org")

        if confidence == "low":
            flags.append("LOW_CONFIDENCE — scores less reliable")

        name = str(row.get("Name", ""))
        if is_key_contact(name, key_contacts):
            flags.append("APA_KEY_CONTACT — in APA Key Contacts list")

        # Private practice check
        pp_keywords = ["private practice", "md llc", "md pllc", "m.d. llc", "psychiatrist.net",
                       "psychiatry associates", "psychiatric associates", "associates llc",
                       "associates pllc", " md ", ", md,"]
        if any(k in inst.lower() for k in pp_keywords) and inst_type == "company":
            flags.append("POSSIBLE_PRIVATE_PRACTICE")

        row_dict = row.to_dict()
        row_dict["FLAGS"] = " | ".join(flags) if flags else ""
        included.append(row_dict)

    # ── Build output dataframe ─────────────────────────────────────────────────
    df_out = pd.DataFrame(included)

    # Sort: clean leads first, then flagged; within each group by score
    df_out["_has_flag"] = df_out["FLAGS"].apply(lambda x: 0 if not x else 1)
    df_out = df_out.sort_values(["_has_flag", "FinalScore"], ascending=[True, False])
    df_out = df_out.drop(columns=["_has_flag"]).reset_index(drop=True)
    df_out.insert(0, "ReviewRank", df_out.index + 1)

    # Keep only the most useful columns for review
    review_cols = [
        "ReviewRank", "FLAGS",
        "Name", "currentTitle", "currentInstitution", "institutionType",
        "Country", "State",
        "FinalScore", "Tier", "confidence",
        "budgetScore", "decisionMakerScore", "isDecisionMaker",
        "financialCapacity", "reasoning",
        "Appearances", "Years",
        "FinalPreScore",
    ]
    # Only keep columns that exist
    review_cols = [c for c in review_cols if c in df_out.columns]
    df_review = df_out[review_cols].copy()

    # ── Save CSV ───────────────────────────────────────────────────────────────
    csv_path  = f"{args.output}_review.csv"
    excl_path = f"{args.output}_excluded.csv"
    df_review.to_csv(csv_path, index=False)
    pd.DataFrame(excluded).to_csv(excl_path, index=False)

    # ── Save Excel with colour coding ──────────────────────────────────────────
    xlsx_path = f"{args.output}_review.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_review.to_excel(writer, sheet_name="Review Leads", index=False)
        ws = writer.sheets["Review Leads"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header row
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        # Row fills
        green_fill  = PatternFill("solid", fgColor="E2EFDA")  # clean lead
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")  # flagged
        alt_green   = PatternFill("solid", fgColor="C6EFCE")
        alt_yellow  = PatternFill("solid", fgColor="FFEB9C")

        green_count = yellow_count = 0
        for row_idx, row_data in enumerate(df_review.itertuples(), start=2):
            flag_val = str(getattr(row_data, "FLAGS", "") or "")
            is_flagged = bool(flag_val.strip())

            if is_flagged:
                fill = alt_yellow if yellow_count % 2 else yellow_fill
                yellow_count += 1
            else:
                fill = alt_green if green_count % 2 else green_fill
                green_count += 1

            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        # Column widths
        col_widths = {
            "ReviewRank": 8, "FLAGS": 45, "Name": 25, "currentTitle": 35,
            "currentInstitution": 35, "institutionType": 14, "Country": 12,
            "State": 7, "FinalScore": 10, "Tier": 6, "confidence": 10,
            "budgetScore": 8, "decisionMakerScore": 8, "isDecisionMaker": 10,
            "financialCapacity": 40, "reasoning": 50,
            "Appearances": 10, "Years": 20, "FinalPreScore": 10,
        }
        for col_idx, col_name in enumerate(review_cols, start=1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = col_widths.get(col_name, 15)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Add a summary sheet
        ws2 = writer.book.create_sheet("Summary")
        clean_count   = len(df_review[df_review["FLAGS"] == ""])
        flagged_count = len(df_review[df_review["FLAGS"] != ""])
        excl_count    = len(excluded)

        excl_reasons = {}
        for e in excluded:
            r = e.get("excludeReason", "Unknown")
            excl_reasons[r] = excl_reasons.get(r, 0) + 1

        summary_rows = [
            ["STEP 2 FILTER SUMMARY", ""],
            ["", ""],
            ["INCLUDED IN REVIEW", ""],
            ["  Clean leads (no flags)", clean_count],
            ["  Flagged leads (review needed)", flagged_count],
            ["  Total for review", clean_count + flagged_count],
            ["", ""],
            ["EXCLUDED", ""],
        ]
        for reason, count in sorted(excl_reasons.items(), key=lambda x: -x[1]):
            summary_rows.append([f"  {reason}", count])
        summary_rows.append(["  TOTAL excluded", excl_count])
        summary_rows.append(["", ""])
        summary_rows.append(["Next step:", f"Review {args.output}_review.xlsx, select leads, then run:"])
        summary_rows.append(["", f"python step3_prepare.py --input {args.output}_review.csv --output {args.output}"])

        for r, row in enumerate(summary_rows, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws2.cell(row=r, column=c, value=val)
                if c == 1 and str(val).isupper():
                    cell.font = Font(bold=True, color="1F4E79")
        ws2.column_dimensions["A"].width = 40
        ws2.column_dimensions["B"].width = 60

    # ── Print summary ──────────────────────────────────────────────────────────
    excl_reasons = {}
    for e in excluded:
        r = e.get("excludeReason", "Unknown").split(" (")[0].split("_")[0]
        excl_reasons[r] = excl_reasons.get(r, 0) + 1

    print(f"""
╔══════════════════════════════════════════╗
║         STEP 2 FILTER COMPLETE           ║
╠══════════════════════════════════════════╣
║  Input records    : {len(df):<20} ║
╠══════════════════════════════════════════╣
║  EXCLUDED                                ║""")
    for reason, count in sorted(excl_reasons.items(), key=lambda x: -x[1]):
        label = f"  {reason}"[:40]
        print(f"║  {label:<38} : {count:<4} ║")
    print(f"""╠══════════════════════════════════════════╣
║  FOR REVIEW (clean)  : {clean_count:<17} ║
║  FOR REVIEW (flagged): {flagged_count:<17} ║
║  TOTAL for review    : {clean_count+flagged_count:<17} ║
╠══════════════════════════════════════════╣
║  Output Excel : {xlsx_path:<24} ║
║  Output CSV   : {csv_path:<24} ║
║  Excluded CSV : {excl_path:<24} ║
╚══════════════════════════════════════════╝

📋 Open {xlsx_path} — green rows are clean, yellow rows need review.
   Delete any rows you don't want to contact, then save as CSV and run:

   python step3_prepare.py --input {csv_path} --output {args.output}
    """)


if __name__ == "__main__":
    main()
