"""
step3_prepare.py
================
Websedge Conference Lead Pipeline — Step 3: Prepare for Ben Leads
------------------------------------------------------------------
Takes your reviewed/selected leads from step2 and prepares them for
copy-paste into the Ben Leads spreadsheet tab.

For each lead:
  1. Serper.dev  → searches for recent news about the institution + official website URL
  2. Mistral AI  → generates all Ben Leads text columns from that context:
       - Focus         (2-4 sentences, what the institution does)
       - Focus Summary (2-4 word specialty tag)
       - Area          (clinical area category)
       - Explanation   (why the institution has financial capacity)
       - Reason/Value/Return from video

Static columns (no AI needed):
  - Affinity/10             based on APA appearance count: 1→5, 2→7, 3→7, 4+→10
  - Affinity Explanation    "X time attendee"
  - Contact/10              always 10  (never contacted)
  - Explanation (Contact)   always "Never Contacted"
  - Potential Lead Overlap? always "No"  (any flags from step2 go into Additional Notes)
  - Source                  "Attendee List"
  - Type                    (BP) Big Private / (MP) Medium Private / (SP) Small Private

Output:
  {output}_ben_leads.xlsx   ← formatted, ready to copy rows into Ben Leads tab
  {output}_ben_leads.csv    ← same as CSV backup

Usage:
  python step3_prepare.py --input apa2026_extra_review.csv --output apa2026_extra

  To run only specific rows, open the review CSV, delete the rows you don't
  want, save, then point --input at your saved file.

API keys:
  MISTRAL_API_KEY  — mistral.ai  (text generation)
  SERPER_API_KEY   — serper.dev  (live institution news + website search)
"""

import argparse, json, os, re, time
import pandas as pd
import requests
from mistralai import Mistral
from datetime import date

# Optional: trafilatura gives much richer context by scraping the institution website.
# Install with:  pip install trafilatura
# If not installed the script still works — falls back to Serper snippets only.
try:
    import trafilatura
    HAS_TRAFILATURA = True
except ImportError:
    HAS_TRAFILATURA = False

# ── Config ─────────────────────────────────────────────────────────────────────

MISTRAL_API_KEY = os.environ.get("MISTRAL_API_KEY", "")
SERPER_API_KEY  = os.environ.get("SERPER_API_KEY",  "")

MISTRAL_MODEL   = "mistral-large-latest"   # best quality for hook writing
SERPER_ENDPOINT = "https://google.serper.dev/search"
BATCH_SIZE      = 5
DELAY           = 1.5   # seconds between Mistral batches

# Domains to skip when picking the official website URL from search results
SOCIAL_DOMAINS = {
    "linkedin.com", "twitter.com", "x.com", "facebook.com",
    "instagram.com", "youtube.com", "wikipedia.org", "indeed.com",
    "glassdoor.com", "yelp.com", "crunchbase.com", "zoominfo.com",
    "bloomberg.com", "pitchbook.com", "doximity.com", "healthgrades.com",
}

# ── Clients ───────────────────────────────────────────────────────────────────

mistral_client = Mistral(api_key=MISTRAL_API_KEY)


# ── Serper: institution news + official website in one call ───────────────────

def serper_search_inst(institution: str) -> tuple:
    """
    One Serper search returns both:
      - news_snippets  (str)  — for Mistral context
      - website_url    (str)  — for the Link column (official site, not social media)
    """
    if not SERPER_API_KEY:
        return ("No news search — SERPER_API_KEY not set.", "")
    query = f'"{institution}" psychiatry mental health 2024 2025 news program'
    try:
        resp = requests.post(
            SERPER_ENDPOINT,
            headers={"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"},
            json={"q": query, "num": 5},
            timeout=10,
        )
        data = resp.json()

        # knowledgeGraph.website is the most reliable official website source
        website = data.get("knowledgeGraph", {}).get("website", "")

        results = data.get("organic", [])[:5]
        lines = []
        for r in results:
            lines.append(f"- {r.get('title', '')}: {r.get('snippet', '')}")
            if not website:
                url = r.get("link", "")
                domain = url.split("/")[2].replace("www.", "") if "://" in url else ""
                if url and not any(sd in domain for sd in SOCIAL_DOMAINS):
                    website = url

        snippets = "\n".join(lines[:4]) if lines else "No recent news found."
        return (snippets, website)

    except Exception as e:
        return (f"Search error: {e}", "")


# ── Website scraper (no API cost — plain HTTP) ────────────────────────────────

def scrape_website(url: str, max_chars: int = 1500) -> str:
    """
    Fetch and extract readable text from an institution's website.
    Returns empty string on any failure — caller falls back to Serper snippets.

    This is a plain HTTP download (no AI, no API cost).
    Requires:  pip install trafilatura
    """
    if not url or not HAS_TRAFILATURA:
        return ""
    try:
        html = trafilatura.fetch_url(url)
        if not html:
            return ""
        text = trafilatura.extract(
            html,
            include_comments=False,
            include_tables=False,
            no_fallback=False,
        )
        return text[:max_chars].strip() if text else ""
    except Exception:
        return ""


# ── Type classification ────────────────────────────────────────────────────────

def classify_type(budget_score: float) -> str:
    """Map budget score → Ben Leads type code."""
    b = float(budget_score or 0)
    if b >= 4: return "(BP) Big Private"
    if b >= 2: return "(MP) Medium Private"
    return "(SP) Small Private"


# ── Affinity score (attendance-based only) ─────────────────────────────────────

def affinity_from_appearances(appearances) -> int:
    """
    APA appearance count → Affinity/10
      1  → 5
      2  → 7
      3  → 7
      4+ → 10
    """
    apps = int(float(appearances or 0))
    if apps >= 4: return 10
    if apps >= 2: return 7
    return 5  # 0 or 1 appearance


# ── CRM paste field ────────────────────────────────────────────────────────────

def paste_in_crm(institution: str, focus: str, tracker_theme: str = "USA Private") -> str:
    """Format the Paste in CRM field exactly as Ben Leads rows do it."""
    focus_snippet = focus[:120].rstrip()
    return (
        f"Call for APA\n"
        f"Tracker Theme: {tracker_theme}\n"
        f"Org: {institution}\n"
        f"Focus: {focus_snippet}"
    )


# ── Mistral: generate all text columns in one call per batch ─────────────────

def generate_columns(leads_ctx: list) -> list:
    """
    Generate Focus, Focus Summary, Area, Explanation, and Reason/Value
    for a list of leads via Mistral.

    Affinity Explanation is now generated statically as "X time attendee"
    and is NOT part of this Mistral call (saves tokens).

    Returns a list of dicts in the same order as leads_ctx.
    """
    all_results = []

    for i in range(0, len(leads_ctx), BATCH_SIZE):
        batch = leads_ctx[i : i + BATCH_SIZE]

        blocks = []
        for j, ctx in enumerate(batch, 1):
            # Prefer scraped website text (institution's own words) over Serper snippets.
            # Fall back to snippets when scraping failed.
            if ctx.get("scraped"):
                institution_context = (
                    f"Institution website content (their own words — use this as primary source):\n"
                    f"{ctx['scraped']}\n\n"
                    f"Recent news snippets (secondary):\n{ctx['news']}"
                )
            else:
                institution_context = f"Recent institution news:\n{ctx['news']}"

            blocks.append(
                f"Person {j}:\n"
                f"Name: {ctx['name']}\n"
                f"Title: {ctx['title']}\n"
                f"Institution: {ctx['institution']}\n"
                f"Institution type: {ctx['inst_type']}\n"
                f"Years attended APA: {ctx['years']}\n"
                f"Number of APA appearances: {ctx['appearances']}\n"
                f"Financial capacity note: {ctx['financial_capacity']}\n"
                f"AI research notes: {ctx['reasoning']}\n"
                f"{institution_context}"
            )

        prompt = (
            "You are filling in a sales lead spreadsheet for Websedge, which produces "
            "$27,500 APA TV documentary features showcased at the American Psychiatric "
            "Association Annual Meeting (25,000+ attendees).\n\n"

            "Study these REAL examples from the spreadsheet so you match the exact style:\n\n"

            "EXAMPLE Focus fields (3-5 sentences, starts lowercase, describes what the "
            "institution does — as if completing 'We were particularly interested in your...'):\n"
            "  • 'focus on personalized treatment plans and telepsychiatry to increase "
            "accessibility, especially your same day access model. We also noted your pathway "
            "teams using clinical advocacy and care navigation to ensure a bespoke care plan "
            "that coordinates housing, mental health, and addiction services for vulnerable "
            "populations'\n"
            "  • 'high-access, technology-driven care designed to eliminate traditional "
            "barriers like waitlists and travel. We also noted your same-day access model, "
            "providing psychiatric evaluations and therapy appointments to address acute "
            "mental health needs without long wait times'\n"
            "  • 'measurement-based care technology platform designed to modernise psychiatric "
            "care through clinical intelligence and integrated workflows'\n\n"

            "EXAMPLE Focus Summary (2-4 words, lowercase):\n"
            "  'addiction treatment', 'telepsychiatry', 'mental health research and care', "
            "'integrated behavioral health'\n\n"

            "EXAMPLE Area (clinical/business category):\n"
            "  'Addiction / Substance Use Disorder', 'Telepsychiatry and Digital Health', "
            "'Child and Adolescent Psychiatry', 'Inpatient Psychiatric Care', "
            "'Correctional Mental Health', 'Geriatric Psychiatry', 'Clinical Research'\n\n"

            "EXAMPLE Explanation (Size-Money — why institution has financial capacity):\n"
            "  '84 employees', 'well-backed, private equity-supported 51-200 employees', "
            "'non-profit health system with ~$2B annual revenue', "
            "'VC-backed digital health startup, Series B funded'\n\n"

            "EXAMPLE Reason/Value/Return from video (what the INSTITUTION gains):\n"
            "  'Boosts psychiatry residency recruitment nationally and showcases their "
            "integrated care model to 25,000+ APA attendees'\n\n"

            "Now generate for each person:\n\n"
            "1. focus — 3-5 sentences, lowercase start, specific to this institution's "
            "actual programs/approach. Reference the recent news where possible. "
            "Each sentence should add new information — don't repeat or pad.\n"
            "2. focusSummary — 2-4 words, lowercase, specialty tag\n"
            "3. area — clinical area category (match the style of the examples above)\n"
            "4. explanation — why the institution has financial capacity (size/revenue/backing)\n"
            "5. videoValue — what the INSTITUTION (not the person) gains from the documentary\n\n"
            f"People:\n\n{'=' * 40}\n" + f"\n{'=' * 40}\n".join(blocks) + "\n\n"
            f"Return a JSON array of exactly {len(batch)} objects. "
            "Keys: focus, focusSummary, area, explanation, videoValue\n"
            "No markdown, no code fences — raw JSON array only."
        )

        for attempt in range(6):
            try:
                resp = mistral_client.chat.complete(
                    model=MISTRAL_MODEL,
                    messages=[
                        {"role": "system", "content":
                            "You write precise, specific content for sales lead spreadsheets. "
                            "Match the exact tone and style of the examples provided. "
                            "Return only valid JSON arrays."},
                        {"role": "user", "content": prompt},
                    ],
                    temperature=0.3,
                    max_tokens=3000,
                )
                text = resp.choices[0].message.content.strip()
                text = re.sub(r"```(?:json)?\s*", "", text).strip().rstrip("`").strip()
                m = re.search(r"(\[.*\])", text, re.DOTALL)
                if m:
                    text = m.group(1)
                parsed = json.loads(text)
                if len(parsed) != len(batch):
                    raise ValueError(f"Expected {len(batch)}, got {len(parsed)}")
                all_results.extend(parsed)
                names = [ctx['name'] for ctx in batch]
                foci  = [r.get('focusSummary', '?') for r in parsed]
                print(f"  ✓ Batch {i//BATCH_SIZE+1}: " +
                      " | ".join(f"{n.split()[0]}: {f}" for n, f in zip(names, foci)))
                break

            except Exception as e:
                wait = 10 * (attempt + 1)
                if attempt < 5:
                    print(f"  ⏳ Attempt {attempt+1} failed ({str(e)[:60]}) — retrying in {wait}s...")
                    time.sleep(wait)
                else:
                    print(f"  ✗ All retries failed for batch {i//BATCH_SIZE+1}. Using placeholders.")
                    for ctx in batch:
                        all_results.append({
                            "focus": f"innovative approach to mental health care at {ctx['institution']}",
                            "focusSummary": "mental health care",
                            "area": "Mental Health",
                            "explanation": ctx.get("financial_capacity", "Financial details not available"),
                            "videoValue": "APA TV documentary provides national visibility to 25,000+ attendees.",
                        })

        time.sleep(DELAY)

    return all_results


# ── Build Ben Leads row ────────────────────────────────────────────────────────

def build_row(lead: dict, gen: dict, ctx: dict) -> dict:
    """Map enriched lead data + generated text into the exact Ben Leads column format."""
    today         = date.today().strftime("%Y-%m-%d")
    institution   = str(lead.get("currentInstitution") or lead.get("Institution", ""))
    title         = str(lead.get("currentTitle") or lead.get("Title", ""))
    name          = str(lead.get("Name", ""))
    country       = str(lead.get("Country", ""))
    state         = str(lead.get("State", ""))
    appearances   = int(float(lead.get("Appearances", 1) or 1))
    years         = str(lead.get("Years", ""))
    budget        = float(lead.get("budgetScore", 0) or 0)
    is_dm         = bool(lead.get("isDecisionMaker", False))
    confidence    = str(lead.get("confidence", "low"))
    tier          = str(lead.get("Tier", ""))
    final_score   = float(lead.get("FinalScore", 0) or 0)
    flags         = str(lead.get("FLAGS", "") or "")
    website       = ctx.get("website", "")

    focus         = gen.get("focus", "")
    focus_summary = gen.get("focusSummary", "")
    area          = gen.get("area", "")
    explanation   = gen.get("explanation", "")
    video_value   = gen.get("videoValue", "")

    size_money   = min(int(budget * 2), 10)
    affinity     = affinity_from_appearances(appearances)
    affinity_exp = f"{appearances} time attendee"

    # Additional Notes: pipeline metadata + any step2 flags
    notes_parts = [f"Tier {tier}", f"Score {final_score:.1f}", f"Confidence: {confidence}"]
    if flags:
        notes_parts.append(f"Flags: {flags}")
    additional_notes = " | ".join(notes_parts)

    return {
        "Date":                           today,
        "Country":                        country,
        "State":                          state,
        "Institution":                    institution,
        "Center Name":                    "",
        "Type":                           classify_type(budget),
        "Position of Lead":               title,
        "Contact Name":                   name,
        "Email":                          "",
        "Tracker Theme":                  "USA Private",
        "Focus":                          focus,
        "Focus Summary":                  focus_summary,
        "Link":                           website,
        "Sent":                           "",
        "Area":                           area,
        "Result":                         "",
        "Pitch Rating":                   "",
        "Source":                         "Attendee List",
        "KWS Term":                       "",
        "Link to Source":                 "",
        "Attended?":                      years,
        "Affinity/10":                    affinity,
        "Affinity Explanation":           affinity_exp,
        "Size-Money/10":                  size_money,
        "Explanation":                    explanation,
        "Contacted? (Year/s)":            "",
        "Contact/10":                     10,
        "Explanation (Contact)":          "Never Contacted",
        "Potential Lead Overlap?":        "No",
        "Reason/Value/Return from video": video_value,
        "Decision Maker?":                "Yes" if is_dm else "Maybe",
        "Additional Notes":               additional_notes,
        "Autocheck Fellows":              0,
        "Paste in CRM":                   paste_in_crm(institution, focus),
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Websedge Step 3 — Prepare Ben Leads")
    parser.add_argument("--input",  required=True,       help="Reviewed CSV from step2 (your selected leads)")
    parser.add_argument("--output", default="prepared",  help="Output filename prefix")
    args = parser.parse_args()

    print(f"\n📂 Loading {args.input}...")
    if args.input.lower().endswith(".xlsx"):
        df = pd.read_excel(args.input, sheet_name="Review Leads")
    else:
        df = pd.read_csv(args.input)
    print(f"   {len(df)} leads to prepare")

    if not SERPER_API_KEY:
        print("   ⚠ SERPER_API_KEY not set — news search disabled. Hooks will be Mistral-only.")
    if not HAS_TRAFILATURA:
        print("   ℹ  trafilatura not installed — website scraping disabled.")
        print("      Run:  pip install trafilatura   to enable richer Focus accuracy.")

    # ── Build context for each lead ───────────────────────────────────────────
    print("\n🔍 Searching institution news + scraping websites via Serper...")
    contexts = []
    for _, row in df.iterrows():
        inst = str(row.get("currentInstitution") or row.get("Institution", ""))

        # 1. Serper: get news snippets + official website URL
        news, website = serper_search_inst(inst)

        # 2. Scrape the website for richer program descriptions (free HTTP — no API)
        #    This is the primary source for Focus accuracy when available.
        scraped = scrape_website(website) if website else ""

        fin = str(row.get("financialCapacity", "") or "")
        if not fin or fin in ("nan", "None"):
            fin = str(row.get("reasoning", "No financial data"))

        contexts.append({
            "name":               str(row.get("Name", "")),
            "title":              str(row.get("currentTitle") or row.get("Title", "")),
            "institution":        inst,
            "inst_type":          str(row.get("institutionType", "unknown")),
            "years":              str(row.get("Years", "")),
            "appearances":        row.get("Appearances", 1),
            "financial_capacity": fin,
            "reasoning":          str(row.get("reasoning", "")),
            "news":               news,
            "website":            website,
            "scraped":            scraped,      # website full text (empty if scrape failed)
        })

        # Status line: show what we got for this lead
        if scraped:
            src = f"✓ scraped ({len(scraped)} chars)"
        elif news and not news.startswith("No recent"):
            src = "~ Serper snippets only"
        else:
            src = "⚠ no source — verify manually"
        print(f"   {inst[:40]:<40} {src}")
        time.sleep(0.3)   # gentle rate limiting

    # ── Generate all text columns via Mistral ─────────────────────────────────
    print(f"\n🤖 Generating Ben Leads columns via Mistral ({MISTRAL_MODEL})...")
    generated = generate_columns(contexts)

    # ── Build output rows ─────────────────────────────────────────────────────
    print("\n📝 Building Ben Leads rows...")
    rows = []
    for lead_row, gen, ctx in zip(df.to_dict(orient="records"), generated, contexts):
        rows.append(build_row(lead_row, gen, ctx))

    df_out = pd.DataFrame(rows)

    # ── Save CSV ───────────────────────────────────────────────────────────────
    csv_path  = f"{args.output}_ben_leads.csv"
    xlsx_path = f"{args.output}_ben_leads.xlsx"
    df_out.to_csv(csv_path, index=False)

    # ── Save Excel matching Ben Leads tab style ────────────────────────────────
    # Build Research Notes dataframe (one row per lead — shows what sources were used)
    research_rows = []
    for ctx in contexts:
        news_text   = ctx.get("news", "")
        scraped     = ctx.get("scraped", "")
        has_news    = news_text not in ("No recent news found.", "") and not news_text.startswith("Search error")
        has_scraped = bool(scraped)

        if has_scraped:
            quality = "✓ Website scraped — highest accuracy"
        elif has_news:
            quality = "~ Serper snippets only — good accuracy"
        else:
            quality = "⚠ No source data — verify Focus manually"

        research_rows.append({
            "Institution":       ctx["institution"],
            "Contact Name":      ctx["name"],
            "Website Found":     ctx.get("website", "") or "—",
            "Source Quality":    quality,
            "Website Content":   scraped[:600] + "…" if len(scraped) > 600 else scraped or "—",
            "Serper Snippets":   news_text,
        })
    df_research = pd.DataFrame(research_rows)

    # Which row indices (0-based) have NO source at all? (for orange highlighting)
    no_news_rows = {
        i for i, ctx in enumerate(contexts)
        if not ctx.get("scraped")
        and (ctx.get("news", "") in ("No recent news found.", "")
             or ctx.get("news", "").startswith("Search error"))
    }

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Ben Leads Ready", index=False)
        df_research.to_excel(writer, sheet_name="Research Notes", index=False)

        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # ── Ben Leads Ready sheet ──────────────────────────────────────────────
        ws = writer.sheets["Ben Leads Ready"]

        # Header row
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=9)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 35

        # Row colours: orange = no news found (verify Focus manually)
        #              alternating white/grey = normal rows
        fill_a      = PatternFill("solid", fgColor="FFFFFF")
        fill_b      = PatternFill("solid", fgColor="F2F2F2")
        fill_nonews = PatternFill("solid", fgColor="FFE0B2")   # pale orange
        for row_idx in range(2, ws.max_row + 1):
            data_idx = row_idx - 2   # 0-based
            if data_idx in no_news_rows:
                fill = fill_nonews
            else:
                fill = fill_a if row_idx % 2 == 0 else fill_b
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

        # Column widths
        col_widths = {
            "Date": 12, "Country": 8, "State": 6, "Institution": 30,
            "Center Name": 20, "Type": 16, "Position of Lead": 30,
            "Contact Name": 22, "Email": 28, "Tracker Theme": 14,
            "Focus": 60, "Focus Summary": 20, "Link": 35, "Sent": 8,
            "Area": 28, "Result": 12, "Pitch Rating": 10, "Source": 14,
            "KWS Term": 10, "Link to Source": 12, "Attended?": 20,
            "Affinity/10": 10, "Affinity Explanation": 20,
            "Size-Money/10": 10, "Explanation": 35,
            "Contacted? (Year/s)": 14, "Contact/10": 10,
            "Explanation (Contact)": 20, "Potential Lead Overlap?": 12,
            "Reason/Value/Return from video": 40,
            "Decision Maker?": 14, "Additional Notes": 35,
            "Autocheck Fellows": 12, "Paste in CRM": 40,
        }
        for col_idx, col_name in enumerate(df_out.columns, start=1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = col_widths.get(col_name, 15)

        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 65

        ws.freeze_panes = "A2"

        # ── Research Notes sheet ───────────────────────────────────────────────
        wr = writer.sheets["Research Notes"]

        # Header
        for cell in wr[1]:
            cell.fill = PatternFill("solid", fgColor="2E7D32")
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        wr.row_dimensions[1].height = 30

        # Rows — highlight no-news rows orange
        for row_idx in range(2, wr.max_row + 1):
            data_idx = row_idx - 2
            row_fill = fill_nonews if data_idx in no_news_rows else fill_a
            for col_idx in range(1, wr.max_column + 1):
                wr.cell(row=row_idx, column=col_idx).fill = row_fill
                wr.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

        wr.column_dimensions["A"].width = 35   # Institution
        wr.column_dimensions["B"].width = 22   # Contact Name
        wr.column_dimensions["C"].width = 40   # Website Found
        wr.column_dimensions["D"].width = 30   # Source Quality
        wr.column_dimensions["E"].width = 60   # Website Content
        wr.column_dimensions["F"].width = 70   # Serper Snippets
        for row_idx in range(2, wr.max_row + 1):
            wr.row_dimensions[row_idx].height = 70

        wr.freeze_panes = "A2"

    scraped_count  = sum(1 for ctx in contexts if ctx.get("scraped"))
    no_source_count = len(no_news_rows)
    print(f"""
╔══════════════════════════════════════════╗
║       STEP 3 PREPARE COMPLETE            ║
╠══════════════════════════════════════════╣
║  Leads prepared  : {len(rows):<21} ║
║  Website scraped : {scraped_count:<21} ║
║  No source found : {no_source_count:<21} ║
║  Output Excel    : {xlsx_path:<21} ║
║  Output CSV      : {csv_path:<21} ║
╚══════════════════════════════════════════╝

📋 Open {xlsx_path}:
   • "Ben Leads Ready" tab  — copy rows into Ben Leads spreadsheet
   • "Research Notes" tab   — shows exactly what was used to write each Focus:
       ✓ Website scraped  = Focus written from institution's own website text
       ~ Serper only      = Focus written from search snippets
       ⚠ Orange rows      = no source found; worth a quick manual check
    """)


if __name__ == "__main__":
    main()
